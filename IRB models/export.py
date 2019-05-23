# xlsx names of regulatory template reporting docs #
    # 'LEICode_PD_ModelID_EndOfObservationPeriod_versionNumber.xlsx'
    # 'LEICode_LGD_ModelID_EndOfObservationPeriod_versionNumber.xlsx'
    # 'LEICode_CCF_ModelID_EndOfObservationPeriod_versionNumber.xlsx'
    # 'LEICode_ELBE_ModelID_EndOfObservationPeriod_versionNumber.xlsx'
    # 'LEICode_LGDD_ModelID_EndOfObservationPeriod_versionNumber.xlsx'
    # 'LEICode_SL_ModelID_EndOfObservationPeriod_versionNumber.xlsx'

#to do: grab the sheet names an loop through the sheets, mind the structure of most files is similat but not the same

import openpyxl
import datetime
import pandas as pd

class export(object):

    def file_path(self, target_filename):
        import os

        local_dr = os.path.normpath(os.path.expanduser("~/Documents/Python"))
        path = local_dr + "/" + target_filename
        return path

    def open_wb(self, file_name):
        """
        As it says
        :param file_name:
        :return:
        """
        return openpyxl.load_workbook(file_name)

    def save_wb(self, file_name):
        """
        As it says
        :param file_name:
        :return:
        """
        return openpyxl.save(file_name)

    def grade_mapping(self, grade_num):
        return {1:"A", 2:"B", 3:"C", 4:"D", 5:"E", 6:"F", 7:"G"}[grade_num]

    def array_toExcel(self, wb, stat_array, row_pos, col_pos, row_wise=True):
        """
        As it says.
        :param wb: excel openpyxl workbook;
        :param stat_array: array to write to excel;
        :param row_pos: initial row positions;
        :param col_pos: column position;
        :param row_wise: True/False.
        :return:
        """
        i = 0
        for stat in stat_array:
            if row_wise:
                wb.cell(row= row_pos + i, column= col_pos).value = stat
                i += 1
            else:
                wb.cell(row= row_pos, column= col_pos + i).value = stat
                i += 1
        return None

    def df_toExcel(self, wb, df, row_pos, col_pos):
        """
        As it says
        :param wb:
        :param df:
        :param row_pos:
        :param col_pos:
        :return:
        """
        for i in range(len(df.index)):
            self.array_toExcel(wb, df.iloc[i,:].values, row_pos + i, col_pos, row_wise = False)
        return None

    def PD_toExcel(self, pd_inputs):
          #Fill PD test statistics and other information to Excel file;
	     #param data_set: development data set
          #param pd_inputs: dictionary containing pd test results, details etc.
          #return: Save to excel results.
          file_name = self.file_path(target_filename=pd_inputs["name"])
          oxl = self.open_wb(file_name)
          # Information missing from test results:
          start_date 				= pd_inputs["start"]
          end_date 					= pd_inputs["end"]
          jeffrey_test				= pd_inputs["jeffrey"].iloc[:-1, :]
          name_rating_grades 			= jeffrey_test.index.tolist()
          nb_rating_grades 			= len(name_rating_grades)
          averagePD_pergrade 			= pd_inputs['avg_PD']
          nb_customer_pergrade 		= pd_inputs['nb_cust']
          nb_default_pergrade 		= jeffrey_test[('Default_Binary', 'sum')].values
          original_exposure_pergrade 	= pd_inputs['orgExp_Grade']
          jeffrey_test_pval_ptf 		= pd_inputs["jeffrey"].iloc[-1, -1]
          nb_customer 				= sum(nb_customer_pergrade)
          # Predictive ability
          ## PD Back-testing using a Jeffreys test (§ 2.5.3.1) - sheet 3.0
          ### Grade Level
          wbk30 = oxl.get_sheet_by_name("3.0")
          self.array_toExcel(wb = wbk30, stat_array = name_rating_grades, row_pos= 8, col_pos= 4)          # Rating grade names
          self.array_toExcel(wb = wbk30, stat_array = averagePD_pergrade, row_pos= 8, col_pos= 5)          # Average PD
          self.array_toExcel(wb = wbk30, stat_array = nb_customer_pergrade, row_pos= 8, col_pos= 6)        # Nb of customers
          self.array_toExcel(wb = wbk30, stat_array = nb_default_pergrade, row_pos= 8, col_pos= 7)         # Nb of defaults
          self.array_toExcel(wb = wbk30, stat_array = jeffrey_test.p_val, row_pos= 8, col_pos= 8)          # Jeffrey p-vals
          self.array_toExcel(wb = wbk30, stat_array = original_exposure_pergrade, row_pos= 8, col_pos= 9)  # Original exposure
          ### Portfolio Level
          wbk30.cell(row = 6, column = 8).value     = jeffrey_test_pval_ptf                                # Jeffrey p-vals at ptf lvl
          # Discriminatory Power
          ## Current AUC vs AUC at initial validation/development (§ 2.5.4.1) - sheet 4.0
          wbk40 = oxl.get_sheet_by_name("4.0")
          self.array_toExcel(wb=wbk40, stat_array = pd_inputs["AUC"], row_pos=7, col_pos=4, row_wise=False)  # AUC
          wbk40.cell(row=7, column=10).value     = start_date                                                # start date
          wbk40.cell(row=7, column=11).value     = end_date                                                  # end date
          wbk40.cell(row=7, column=12).value     = nb_customer                                               # nb of customers
          wbk40.cell(row=7, column=13).value     = pd_inputs["AUC_init"]                                     # nb of customers
          wbk40.cell(row=18, column=8).value     = start_date                                                # nb of customers
          wbk40.cell(row=18, column=9).value     = end_date                                                  # nb of customers
          wbk40.cell(row=18, column=10).value    = nb_customer                                               # nb of customers
          wbk40.cell(row=18, column=11).value    = nb_rating_grades                                          # nb of customers
          # Stability - Customer Migration
          ## Current AUC vs AUC at initial validation/development (§ 2.5.4.1) - sheet 4.0
          self.array_toExcel(wb=wbk40, stat_array=pd_inputs["concentration_rating_grades"], row_pos=18, col_pos=4, row_wise=False)
          ## Customer Migrations (§ 2.5.5.1) - sheet 5.1
          wbk51 = oxl.get_sheet_by_name("5.1")
          self.array_toExcel(wb=wbk51, stat_array=pd_inputs["customer_migrations"], row_pos=7, col_pos=4,row_wise=False)
          # Stability - Stability of Migrations
          ## Customer Migrations (§ 2.5.5.2) - sheet 5.2
          wbk52 = oxl.get_sheet_by_name("5.2")
          c = 0
          for j in range(len(pd_inputs["stability_migration_matrix"][0].columns) - 3):
               self.array_toExcel(wb=wbk52, stat_array = pd_inputs["stability_migration_matrix"][0].iloc[:, j],
                                  row_pos=7, col_pos=(4 + c))      # transition probability
               self.array_toExcel(wb=wbk52, stat_array = pd_inputs["stability_migration_matrix"][1][:, j],
                                  row_pos=7, col_pos=(5 + c))      # z
               self.array_toExcel(wb=wbk52, stat_array = pd_inputs["stability_migration_matrix"][2][:, j], row_pos=7,
                                  col_pos=(6 + c))           # phi
               c += 3

          self.array_toExcel(wb=wbk52, stat_array = pd_inputs["stability_migration_matrix"][0].iloc[:,-1:], row_pos=7,
                             col_pos=124) #freq of dflt customers
          self.df_toExcel(wb=wbk52, df=pd_inputs["stability_migration_matrix"][0].iloc[:,-3:], row_pos=7, col_pos=124)

          # Save file
          oxl.save( file_name )
          oxl.close()
          return "PD results saved to Excel."

    def LGD_toExcel(self, data_set, lgd_inputs):
        """
        Fill LGD stats and other informations.
        :param self:
        :param data_set:
        :param lgd_inputs:
        :return:
        """
        file_name = self.file_path(target_filename="LEICode_LGD_ModelID_EndOfObservationPeriod_versionNumber.xlsx")
        oxl = openpyxl.load_workbook(file_name)

        # Information missing from test results:
        start_date = datetime.date(2007, 1, 1)
        end_date = datetime.date(2015, 1, 1)
        nb_customer = len(data_set.id.unique())
        grade_nb = data_set[data_set.Default_Binary == 1].Bin_LGD.unique() #nb of estimated LGD bins/grades
        grade_name = []
        avLGDE_perGrade = []
        avLGDR_perGrade = []
        for g in range(1, len(grade_nb) + 1):
            grade_name.append( self.grade_mapping(grade_num = g) )
            avLGDE_perGrade.append( data_set.groupby("Bin_LGD").LGD.mean()[g] )
            avLGDR_perGrade.append( data_set.groupby("Bin_LGD").LGD_realised.mean()[g] )

        avLGDE = data_set.LGD.mean()
        avLGDR = data_set.LGD_realised.mean()

        # Construct contingency table for sheet 2.1: frequency of realised LGD per bucket/grade of estimated LGD:
        rg = [0.0] + avLGDE_perGrade + [1.0]
        ctgy_perGrade = data_set[data_set.Default_Binary == 1].groupby(["Bin_LGD", pd.cut(data_set[data_set.Default_Binary == 1].LGD_realised, rg)]).LGD_realised.count().unstack()
        ctgy_ptf = data_set[data_set.Default_Binary == 1].groupby([pd.cut(data_set[data_set.Default_Binary == 1].LGD_realised, rg)]).LGD_realised.count().values
        grade_counts = ctgy_perGrade.sum(axis=1).values
        nb_facilities = sum(grade_counts)


        # Predictive ability
        ## LGD back-testing using a t-test (§ 2.6.2.1) - sheet 2.0 or 2.1
        if len( data_set.Bin_LGD.unique() ) > 20:
            wbk2 = oxl.get_sheet_by_name("2.0")
            col_start = 5
            col_end = 21
        else:
            wbk2 = oxl.get_sheet_by_name("2.1")
            col_start = 4
            col_end = 30
        # Grade Level
        self.array_toExcel(wb=wbk2, stat_array= grade_name, row_pos=9, col_pos=col_start, row_wise=True)
        self.array_toExcel(wb=wbk2, stat_array= grade_counts, row_pos=9, col_pos=col_start + 1, row_wise=True)
        self.array_toExcel(wb=wbk2, stat_array= avLGDE_perGrade, row_pos=9, col_pos=7, row_wise=True)
        self.array_toExcel(wb=wbk2, stat_array= avLGDR_perGrade, row_pos=9, col_pos=8, row_wise=True)
        self.df_toExcel(wb=wbk2, df= ctgy_perGrade, row_pos=9, col_pos=9) #Contingency table
        self.df_toExcel(wb=wbk2, df= pd.DataFrame(lgd_inputs["predictive_ability"][1]).T, row_pos=9, col_pos=col_end) #per grade

        # Ptf lvl
        self.array_toExcel(wb=wbk2, stat_array=lgd_inputs["predictive_ability"][0], row_pos=7, col_pos=col_end,
                           row_wise=False)
        self.array_toExcel(wb=wbk2, stat_array=ctgy_ptf, row_pos=7, col_pos=9, row_wise=False)
        wbk2.cell(row=7, column=5).value = nb_facilities # Number of facilities
        wbk2.cell(row=7, column=6).value = 0.0  # Number-weighted average of estimated LGD without downturn component(if available)
        wbk2.cell(row=7, column=7).value = avLGDE # Number-weighted average of estimated LGD
        wbk2.cell(row=7, column=8).value = avLGDR # Number-weighted average of realised LGD

        # Discriminatory Power
        ## Current gAUC vs gAUC at initial validation/development (§ 2.6.3.1) - sheet 3.0
        wbk30 = oxl.get_sheet_by_name("3.0")
        self.array_toExcel(wb=wbk30, stat_array=lgd_inputs["AUC"][:-1], row_pos=7, col_pos=4, row_wise=False)
        wbk30.cell(row= 7, column= 9).value = start_date # start date
        wbk30.cell(row=7, column=10).value = end_date # end date
        wbk30.cell(row=7, column=11).value = nb_customer # nb of customers
        wbk30.cell(row=7, column=12).value = lgd_inputs["AUC"][-1] # Variance (gAUC_init)

        # Save file
        oxl.save(file_name)
        oxl.close()
        return "LGD results saved to Excel."

    def CCF_toExcel(self, data_set, ccf_inputs):
        """
        Fill CCF stats and other informations.
        :param self:
        :param data_set:
        :param lgd_inputs:
        :return:
        """
        file_name   = self.file_path(target_filename="LEICode_CCF_ModelID_EndOfObservationPeriod_versionNumber.xlsx")
        oxl         = openpyxl.load_workbook(file_name)

        # Information missing from test results:
        start_date	        = datetime.date(2007, 1, 1)
        end_date	        = datetime.date(2015, 1, 1)
        nb_customer         = len(data_set.id.unique())
        grade_nb            = data_set.Bin_CCF.unique()
        grade_name          = []
        grade_counts        = []
        avCCFE_perGrade     = []
        avCCFR_perGrade     = []
        minCCFR_perGrade    = []
        maxCCFR_perGrade    = []
        q5CCFR_perGrade     = []
        q10CCFR_perGrade    = []
        q25CCFR_perGrade    = []
        q50CCFR_perGrade    = []
        q75CCFR_perGrade    = []
        q90CCFR_perGrade    = []
        q95CCFR_perGrade    = []
        for g in range(1, len(grade_nb) + 1):
            grade_name.append( self.grade_mapping(grade_num = g) )
            grade_counts.append( data_set[data_set.Default_Binary == 1]["Bin_CCF"].value_counts()[g] )
            avCCFE_perGrade.append( data_set.groupby("Bin_CCF").CCF.mean()[g] )
            avCCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.mean()[g] )
            minCCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.min()[g])
            maxCCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.max()[g])
            q5CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.05)[g])
            q10CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.10)[g])
            q25CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.25)[g])
            q50CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.50)[g])
            q75CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.75)[g])
            q90CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.90)[g])
            q95CCFR_perGrade.append( data_set.groupby("Bin_CCF").CCF_realised.quantile(0.95)[g])

        bcktesting_ccf_ptf = ["N/A", #Name of facility grade/pool or segment
                              len(data_set.id.unique()), # Number of facilities (R)
                              data_set.CCF.mean(), # Average estimated CCF (CCF^E)
                              data_set.CCF_realised.mean(), # Average realised CCF (CCF^R)
                              0.0, # Floor used (if applicable)
                              0.0, # Number of CCF realisations floored
                              data_set.CCF_realised.min(), # Minimum CCF^R
                              data_set.CCF_realised.quantile(0.05), # Quantiles
                              data_set.CCF_realised.quantile(0.10), #
                              data_set.CCF_realised.quantile(0.25), #
                              data_set.CCF_realised.quantile(0.50), #
                              data_set.CCF_realised.quantile(0.75), #
                              data_set.CCF_realised.quantile(0.90), #
                              data_set.CCF_realised.quantile(0.95), #
                              data_set.CCF_realised.max(),          # Maximum CCF^R
                              0 # Exposure-weighted average of CCF^R (to be created)
                              ]

        # Predictive ability
        ## CCF back-testing using a t-test (§ 2.9.3.1) - sheet 3.1
        wbk31 = oxl.get_sheet_by_name("3.1")
        # Grade Lvl
        self.array_toExcel(wb=wbk31, stat_array = grade_name, row_pos=10, col_pos=4, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array = grade_counts, row_pos=10, col_pos=5, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array = avCCFE_perGrade, row_pos=10, col_pos=6, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array = avCCFR_perGrade, row_pos=10, col_pos=7, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array=[0] * 7, row_pos=10, col_pos=8, row_wise=True) # Floor used (if applicable)
        self.array_toExcel(wb=wbk31, stat_array=[0] * 7, row_pos=10, col_pos=9, row_wise=True) # Number of CCF realisations floored
        self.array_toExcel(wb=wbk31, stat_array= minCCFR_perGrade, row_pos=10, col_pos=10, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= maxCCFR_perGrade, row_pos=10, col_pos=18, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array=[0] * 7, row_pos=10, col_pos=19, row_wise=True) # Exposure-weighted average of CCF^R (to be created)
        self.array_toExcel(wb=wbk31, stat_array= q5CCFR_perGrade, row_pos=10, col_pos=11, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q10CCFR_perGrade, row_pos=10, col_pos=12, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q25CCFR_perGrade, row_pos=10, col_pos=13, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q50CCFR_perGrade, row_pos=10, col_pos=14, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q75CCFR_perGrade, row_pos=10, col_pos=15, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q90CCFR_perGrade, row_pos=10, col_pos=16, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= q95CCFR_perGrade, row_pos=10, col_pos=17, row_wise=True)
        self.array_toExcel(wb=wbk31, stat_array= [0] * 7, row_pos=10, col_pos=23, row_wise=True) # Number of facilities excluded due to outlier handling (set to zero)

        # Ptf Lvl
        self.df_toExcel(wb=wbk31, df = pd.DataFrame(ccf_inputs["predictive_ability"][1]).T, row_pos=10, col_pos=20)
        self.array_toExcel(wb=wbk31, stat_array=ccf_inputs["predictive_ability"][0], row_pos=8, col_pos=20, row_wise=False)
        self.array_toExcel(wb=wbk31, stat_array=bcktesting_ccf_ptf, row_pos=8, col_pos=4, row_wise=False)
        wbk31.cell(row=8, column=23).value = 0  # Number of facilities excluded due to outlier handling

        # Discriminatory Power
        ## Current gAUC vs gAUC at initial validation/development (§ 2.9.3.1) - sheet 4.0
        wbk40 = oxl.get_sheet_by_name("4.0")
        self.array_toExcel(wb=wbk40, stat_array=ccf_inputs["AUC"][:-1], row_pos=7, col_pos=4, row_wise=False)
        wbk40.cell(row= 7, column= 10).value = start_date # start date
        wbk40.cell(row=7, column=11).value = end_date # end date
        wbk40.cell(row=7, column=12).value = nb_customer # nb of customers
        wbk40.cell(row=7, column=13).value = ccf_inputs["AUC"][-1] # Variance (gAUC_init)

        # Save file
        oxl.save(file_name)
        oxl.close()
        return "CCF results saved to Excel."
